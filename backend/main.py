"""Reviews Intelligence — FastAPI app."""

import asyncio
import logging
import os
import secrets
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, List, Optional, Set

from fastapi import (
    BackgroundTasks, Depends, FastAPI, HTTPException,
    Request, Response, WebSocket, WebSocketDisconnect,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy import desc, select
from starlette.middleware.sessions import SessionMiddleware

from .auth import get_google_auth_url, get_gmail_auth_url, handle_oauth_callback, require_auth
from .database import AsyncSessionLocal, ChromeJob, EmailSend, ScanRun, User, init_db
from .scanner.engine import run_scan
from .scanner.branding import fetch_brand_logo
from .sf_client import search_sf_accounts
from .sf_contacts import get_contacts_for_brand
from .gmail_client import send_email as gmail_send
from .chrome_converter import ChromeAuditData, chrome_data_to_signals, score_from_chrome_data
from .chrome_processor import chrome_job_processor, queue_chrome_job

log = logging.getLogger("main")

app = FastAPI(title="Reviews Intelligence", docs_url=None, redoc_url=None)

app.add_middleware(
    SessionMiddleware,
    secret_key=os.environ.get("SECRET_KEY", secrets.token_urlsafe(32)),
    session_cookie="ri_session",
    https_only=bool(os.environ.get("RAILWAY_PUBLIC_DOMAIN") or os.environ.get("NODE_ENV") == "production"),
    same_site="lax",
    max_age=8 * 60 * 60,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── WebSocket connection manager ──────────────────────────────────────────────

class WSManager:
    def __init__(self):
        self._connections: Dict[str, List[WebSocket]] = {}

    async def connect(self, scan_id: str, ws: WebSocket):
        await ws.accept()
        self._connections.setdefault(scan_id, []).append(ws)

    def disconnect(self, scan_id: str, ws: WebSocket):
        conns = self._connections.get(scan_id, [])
        if ws in conns:
            conns.remove(ws)

    async def broadcast(self, scan_id: str, message: dict):
        for ws in list(self._connections.get(scan_id, [])):
            try:
                await ws.send_json(message)
            except Exception:
                self.disconnect(scan_id, ws)

ws_manager = WSManager()
runner_last_seen_at: Optional[datetime] = None


def _runner_status() -> dict:
    if not runner_last_seen_at:
        return {"online": False, "last_seen_at": None, "seconds_since_seen": None}
    seconds = max(0, int((datetime.utcnow() - runner_last_seen_at).total_seconds()))
    return {
        "online": seconds <= 30,
        "last_seen_at": _iso(runner_last_seen_at),
        "seconds_since_seen": seconds,
    }


# ── Startup ───────────────────────────────────────────────────────────────────

@app.on_event("startup")
async def startup():
    await init_db()
    log.info("DB initialized.")
    # Start Chrome job processor as background task
    asyncio.create_task(chrome_job_processor())
    log.info("Chrome job processor started.")


# ── Auth routes ───────────────────────────────────────────────────────────────

@app.get("/auth/login")
async def auth_login(request: Request):
    url = get_google_auth_url(request)
    return RedirectResponse(url)


@app.get("/auth/callback")
async def auth_callback(request: Request, code: str = "", state: str = "", error: str = ""):
    if error:
        return RedirectResponse(f"/login?error={error}")
    if not code:
        return RedirectResponse("/login?error=no_code")
    oauth_type = request.session.get("oauth_type", "login")
    return_to = request.session.pop("oauth_return_to", "/")
    try:
        async with AsyncSessionLocal() as db:
            user = await handle_oauth_callback(code, state, request, db)
        request.session["user"] = {
            "id": str(user.id),
            "email": user.email,
            "name": user.name,
            "photo": user.profile_photo,
            "gmail_connected": bool(user.gmail_refresh_token),
        }
        if oauth_type == "gmail":
            return RedirectResponse(return_to + "?gmail=connected")
        return RedirectResponse("/")
    except HTTPException as e:
        if e.status_code == 403:
            return RedirectResponse("/login?error=domain")
        return RedirectResponse(f"/login?error=oauth")


@app.get("/auth/gmail")
async def auth_gmail(request: Request, return_to: str = "/", user: dict = Depends(require_auth)):
    """Start Gmail OAuth flow to get gmail.send permission + refresh token."""
    request.session["oauth_return_to"] = return_to
    url = get_gmail_auth_url(request)
    return RedirectResponse(url)


@app.post("/auth/logout")
async def auth_logout(request: Request):
    request.session.clear()
    return JSONResponse({"ok": True})


# ── Health check ──────────────────────────────────────────────────────────────

@app.get("/health")
async def health():
    return {"status": "ok"}


# ── API routes ─────────────────────────────────────────────────────────────────

@app.get("/api/me")
async def api_me(request: Request, user: dict = Depends(require_auth)):
    # Enrich with live gmail status from DB
    async with AsyncSessionLocal() as db:
        result = await db.execute(select(User).where(User.email == user["email"]))
        db_user = result.scalar_one_or_none()
    gmail_connected = bool(db_user and db_user.gmail_refresh_token)
    return {**user, "gmail_connected": gmail_connected}


@app.get("/api/sf-accounts")
async def api_sf_accounts(q: str = "", user: dict = Depends(require_auth)):
    return search_sf_accounts(q)


@app.get("/api/scans")
async def api_list_scans(limit: int = 20, user: dict = Depends(require_auth)):
    async with AsyncSessionLocal() as db:
        result = await db.execute(
            select(ScanRun).order_by(desc(ScanRun.triggered_at)).limit(limit)
        )
        runs = result.scalars().all()
    return [_serialize_run(r) for r in runs]


class ScanRequest(BaseModel):
    domain: str
    brand_name: str
    account_owner: Optional[str] = "Unknown"
    sf_reviews_provider: Optional[str] = None
    sf_loyalty_provider: Optional[str] = None


@app.post("/api/scans")
async def api_create_scan(
    body: ScanRequest,
    background_tasks: BackgroundTasks,
    user: dict = Depends(require_auth),
):
    scan_id = str(uuid.uuid4())
    async with AsyncSessionLocal() as db:
        run = ScanRun(
            id=scan_id,
            brand_name=body.brand_name,
            domain=body.domain,
            triggered_by=user["email"],
            status="pending",
            sf_platform=body.sf_reviews_provider,
        )
        db.add(run)
        await db.commit()

    async def _run():
        await run_scan(
            scan_id=scan_id,
            brand_name=body.brand_name,
            domain=body.domain,
            account_owner=body.account_owner or user["name"],
            sf_reviews_provider=body.sf_reviews_provider,
            triggered_by=user["email"],
            broadcast=ws_manager.broadcast,
        )

    background_tasks.add_task(_run)
    return {"scan_id": scan_id}


@app.get("/api/scans/{scan_id}")
async def api_get_scan(scan_id: str, user: dict = Depends(require_auth)):
    async with AsyncSessionLocal() as db:
        run = await db.get(ScanRun, scan_id)
    if not run:
        raise HTTPException(404, "Scan not found")
    return _serialize_run(run)


@app.get("/api/scans/{scan_id}/pdf")
async def api_download_pdf(scan_id: str, user: dict = Depends(require_auth)):
    """Serve PDF from disk if available, else regenerate via Playwright."""
    import re as _re
    from fastapi.responses import Response as FastAPIResponse
    from .scanner.pdf_generator import render_html
    from .scanner.browser import generate_pdf_bytes

    async with AsyncSessionLocal() as db:
        run = await db.get(ScanRun, scan_id)
    if not run:
        raise HTTPException(404, "Scan not found")
    if run.status != "complete":
        raise HTTPException(404, "Scan not complete yet")

    safe = _re.sub(r"[^\w]", "_", run.brand_name)
    filename = f"{safe}_audit.pdf"

    # Try to serve previously generated file
    pdf_dir = Path(os.environ.get("PDF_DIR", "/pdfs"))
    pdf_file = pdf_dir / f"{scan_id}.pdf"
    if pdf_file.exists() and pdf_file.stat().st_size > 1000:
        log.info("Serving cached PDF for scan %s (%d bytes)", scan_id, pdf_file.stat().st_size)
        return FileResponse(
            str(pdf_file),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # Regenerate from DB data. Pass items through as-is — evidence items are
    # {path, caption} dicts the PDF renders with their finding captions.
    screenshot_paths = [
        item for item in (run.screenshots_json or [])
        if (isinstance(item, str)) or (isinstance(item, dict) and item.get("path"))
    ]

    html = render_html(
        brand_name=run.brand_name,
        domain=run.domain,
        account_owner=run.triggered_by or "",
        overall_score=run.overall_score or 0,
        grade=run.grade or "D",
        scores=run.scores_json or {},
        pitch_angles=run.pitch_angles_json or [],
        detected_platform=run.detected_platform,
        sf_platform=run.sf_platform,
        platform_mismatch=run.platform_mismatch or False,
        vertical=None,
        page_speed_score=None,
        page_speed_lcp="",
        screenshot_paths=screenshot_paths,
        brand_logo_b64=await fetch_brand_logo(run.domain),
        scan_ts=run.triggered_at.isoformat() if run.triggered_at else "",
    )

    try:
        log.info("Generating PDF via Playwright for scan %s", scan_id)
        pdf_bytes = await generate_pdf_bytes(html)
        log.info("PDF generated: %d bytes for scan %s", len(pdf_bytes), scan_id)
        # Cache to disk for next request
        try:
            pdf_dir.mkdir(parents=True, exist_ok=True)
            pdf_file.write_bytes(pdf_bytes)
        except Exception:
            pass
        return FastAPIResponse(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as exc:
        log.error("Playwright PDF failed for scan %s: %s", scan_id, exc, exc_info=True)
        # Fallback: return HTML so user can print → Save as PDF
        return FastAPIResponse(
            content=html,
            media_type="text/html; charset=utf-8",
            headers={"Content-Disposition": f'inline; filename="{filename}.html"'},
        )


@app.get("/api/scans/{scan_id}/logs")
async def api_scan_logs(scan_id: str, user: dict = Depends(require_auth)):
    """Return the step-by-step audit log for a scan (for debugging)."""
    async with AsyncSessionLocal() as db:
        run = await db.get(ScanRun, scan_id)
    if not run:
        raise HTTPException(404, "Scan not found")
    return {
        "scan_id": scan_id,
        "brand_name": run.brand_name,
        "status": run.status,
        "audit_log": run.audit_log_json or [],
    }


# ── Contacts + Email Send ─────────────────────────────────────────────────────

@app.get("/api/scans/{scan_id}/contacts")
async def api_scan_contacts(scan_id: str, user: dict = Depends(require_auth)):
    """Return SF contacts for the brand being scanned."""
    async with AsyncSessionLocal() as db:
        run = await db.get(ScanRun, scan_id)
    if not run:
        raise HTTPException(404, "Scan not found")
    contacts = get_contacts_for_brand(run.brand_name, run.domain)
    return {"brand_name": run.brand_name, "domain": run.domain, "contacts": contacts}


class SendEmailRequest(BaseModel):
    contacts: List[dict]          # [{email, first_name, last_name, title}]
    subject: str
    body: str                     # may contain {{first_name}} placeholder


@app.post("/api/scans/{scan_id}/send-emails")
async def api_send_emails(
    scan_id: str,
    payload: SendEmailRequest,
    user: dict = Depends(require_auth),
):
    """Send personalised Slinger emails via Gmail to selected contacts."""
    # Look up the user's Gmail refresh token
    async with AsyncSessionLocal() as db:
        result = await db.execute(select(User).where(User.email == user["email"]))
        db_user = result.scalar_one_or_none()

    if not db_user or not db_user.gmail_refresh_token:
        raise HTTPException(403, "Gmail not connected — visit /auth/gmail to connect")

    run = None
    async with AsyncSessionLocal() as db:
        run = await db.get(ScanRun, scan_id)
    if not run:
        raise HTTPException(404, "Scan not found")

    results = []
    for contact in payload.contacts:
        to_email = contact.get("email", "")
        first_name = contact.get("first_name", "").strip() or contact.get("last_name", "").strip() or "there"
        if not to_email:
            continue

        # Personalise body — replace {{first_name}} placeholder
        personalised_body = payload.body.replace("{{first_name}}", first_name)

        send_result = await gmail_send(
            refresh_token=db_user.gmail_refresh_token,
            to_email=to_email,
            subject=payload.subject,
            body=personalised_body,
        )

        # Log to DB
        async with AsyncSessionLocal() as db:
            log_row = EmailSend(
                scan_id=run.id,
                sent_by=user["email"],
                to_email=to_email,
                to_name=f"{contact.get('first_name','')} {contact.get('last_name','')}".strip(),
                subject=payload.subject,
                body=personalised_body,
                gmail_message_id=send_result.get("message_id"),
                status="sent" if send_result["ok"] else "failed",
                error=send_result.get("error"),
            )
            db.add(log_row)
            await db.commit()

        results.append({
            "email": to_email,
            "name": f"{contact.get('first_name','')} {contact.get('last_name','')}".strip(),
            **send_result,
        })

    sent = sum(1 for r in results if r["ok"])
    failed = len(results) - sent
    return {"sent": sent, "failed": failed, "results": results}


@app.get("/api/scans/{scan_id}/chrome-status")
async def api_chrome_status(scan_id: str, user: dict = Depends(require_auth)):
    """Frontend polls this while Chrome is running."""
    async with AsyncSessionLocal() as db:
        run = await db.get(ScanRun, scan_id)
    if not run:
        raise HTTPException(404, "Scan not found")
    return {
        "scan_id": scan_id,
        "scan_mode": run.scan_mode,
        "chrome_job_status": run.chrome_job_status,
        "chrome_job_queued_at": _iso(run.chrome_job_queued_at),
        "chrome_job_started_at": _iso(run.chrome_job_started_at),
        "chrome_job_completed_at": _iso(run.chrome_job_completed_at),
        "chrome_pdps_visited": run.chrome_pdps_visited or 0,
        "chrome_error": run.chrome_error,
        "scan_fallback_reason": run.scan_fallback_reason,
        "overall_status": run.status,
        "runner": _runner_status(),
    }


class ChromeFallbackRequest(BaseModel):
    reason: str = "manual"


class ChromeJobFailureRequest(BaseModel):
    error: str = "Browser Scan failed"
    blocked: bool = False


@app.post("/api/scans/{scan_id}/chrome-fallback")
async def api_chrome_fallback(
    scan_id: str,
    body: ChromeFallbackRequest,
    user: dict = Depends(require_auth),
):
    """Manually trigger a Chrome fallback for any completed or failed scan."""
    async with AsyncSessionLocal() as db:
        run = await db.get(ScanRun, scan_id)
    if not run:
        raise HTTPException(404, "Scan not found")

    from .scanner.utils import domain_to_url
    base_url = domain_to_url(run.domain)
    job_id = await queue_chrome_job(
        scan_id=scan_id,
        brand_name=run.brand_name,
        domain=run.domain,
        base_url=base_url,
        fallback_reason=body.reason,
        priority=2,  # manual triggers get higher priority
    )

    # Count position in queue
    async with AsyncSessionLocal() as db:
        from sqlalchemy import select, func
        result = await db.execute(
            select(func.count()).select_from(ChromeJob).where(ChromeJob.status == "queued")
        )
        position = result.scalar() or 1

    return {"job_id": job_id, "position": int(position)}


@app.get("/api/chrome-queue")
async def api_chrome_queue(user: dict = Depends(require_auth)):
    """Returns the current Chrome job queue status."""
    from sqlalchemy import select, func
    from datetime import date

    async with AsyncSessionLocal() as db:
        # Currently running
        running_result = await db.execute(
            select(ChromeJob).where(ChromeJob.status == "running").limit(1)
        )
        running_job = running_result.scalar_one_or_none()

        # Queued jobs
        queued_result = await db.execute(
            select(ChromeJob)
            .where(ChromeJob.status == "queued")
            .order_by(ChromeJob.priority.desc(), ChromeJob.created_at.asc())
        )
        queued_jobs = queued_result.scalars().all()

        # Completed today
        today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        completed_result = await db.execute(
            select(func.count()).select_from(ChromeJob)
            .where(ChromeJob.status == "complete", ChromeJob.completed_at >= today_start)
        )
        completed_today = completed_result.scalar() or 0

    return {
        "running": {
            "scan_id": str(running_job.scan_id),
            "brand": running_job.brand_name,
            "started_at": _iso(running_job.started_at),
        } if running_job else None,
        "queued": [
            {
                "scan_id": str(j.scan_id),
                "brand": j.brand_name,
                "position": i + 1,
                "queued_at": _iso(j.created_at),
            }
            for i, j in enumerate(queued_jobs)
        ],
        "completed_today": completed_today,
        "runner": _runner_status(),
    }


@app.post("/api/chrome-jobs/next")
async def api_claim_chrome_job(request: Request):
    """Claim the next queued Browser Scan job (pulled by chrome_runner.py).

    Authenticated with the webhook secret (the runner has no user session).
    Also reclaims jobs whose runner died mid-scan: past their timeout they are
    requeued, or failed once max attempts are exhausted.
    """
    secret = request.headers.get("X-Webhook-Secret", "")
    expected = os.environ.get("BROWSER_WEBHOOK_SECRET", "")
    if not expected or secret != expected:
        raise HTTPException(status_code=401, detail="Invalid webhook secret")

    global runner_last_seen_at
    runner_last_seen_at = datetime.utcnow()

    now = datetime.utcnow()
    async with AsyncSessionLocal() as db:
        # Reclaim stale 'running' jobs (runner crashed / lost connection)
        stale = await db.execute(
            select(ChromeJob).where(
                ChromeJob.status == "running", ChromeJob.timeout_at <= now
            )
        )
        for j in stale.scalars().all():
            if (j.attempts or 0) >= (j.max_attempts or 2):
                j.status = "failed"
                j.error = "Exceeded max attempts"
                j.completed_at = now
                run = await db.get(ScanRun, j.scan_id)
                if run:
                    run.chrome_job_status = "failed"
                    run.chrome_error = j.error
                    run.chrome_job_completed_at = now
            else:
                j.status = "queued"
        await db.commit()

        # Claim the next queued job (highest priority, oldest first). Row locking
        # prevents two local runners from claiming the same Browser Scan.
        result = await db.execute(
            select(ChromeJob)
            .where(ChromeJob.status == "queued")
            .order_by(ChromeJob.priority.desc(), ChromeJob.created_at.asc())
            .with_for_update(skip_locked=True)
            .limit(1)
        )
        job = result.scalar_one_or_none()
        if not job:
            return {"job": None}

        job.status = "running"
        job.started_at = now
        job.timeout_at = now + timedelta(minutes=15)
        job.attempts = (job.attempts or 0) + 1
        run = await db.get(ScanRun, job.scan_id)
        if run:
            run.chrome_job_status = "running"
            run.chrome_job_started_at = now
        await db.commit()

        return {"job": {
            "id": str(job.id),
            "scan_id": str(job.scan_id),
            "brand": job.brand_name,
            "domain": job.domain,
            "base_url": job.base_url,
            "attempts": job.attempts,
        }}


@app.post("/api/chrome-jobs/{job_id}/fail")
async def api_fail_chrome_job(job_id: str, body: ChromeJobFailureRequest, request: Request):
    """Mark a Browser Scan job failed immediately instead of waiting for timeout."""
    secret = request.headers.get("X-Webhook-Secret", "")
    expected = os.environ.get("BROWSER_WEBHOOK_SECRET", "")
    if not expected or secret != expected:
        raise HTTPException(status_code=401, detail="Invalid webhook secret")

    now = datetime.utcnow()
    async with AsyncSessionLocal() as db:
        job = await db.get(ChromeJob, job_id)
        if not job:
            raise HTTPException(404, "Chrome job not found")
        job.status = "failed"
        job.error = body.error[:1000]
        job.completed_at = now
        run = await db.get(ScanRun, job.scan_id)
        if run:
            run.chrome_job_status = "failed"
            run.chrome_error = body.error[:1000]
            run.chrome_job_completed_at = now
            if body.blocked and run.status != "complete":
                run.status = "blocked"
                run.error_message = body.error[:1000]
        await db.commit()
    return {"ok": True}


@app.post("/api/browser-data/{scan_id}")
async def receive_browser_data(
    scan_id: str,
    data: ChromeAuditData,
    request: Request,
):
    """
    Webhook endpoint: Claude in Chrome POSTs completed audit data here.
    Validates secret, scores all dimensions, generates PDF and Slinger drafts,
    saves everything to DB, notifies frontend via WebSocket.
    """
    # Validate webhook secret
    secret = request.headers.get("X-Webhook-Secret", "")
    expected = os.environ.get("BROWSER_WEBHOOK_SECRET", "")
    if not expected or secret != expected:
        log.warning("Webhook auth failure for scan %s (secret mismatch)", scan_id)
        raise HTTPException(status_code=401, detail="Invalid webhook secret")

    async with AsyncSessionLocal() as db:
        run = await db.get(ScanRun, scan_id)
    if not run:
        raise HTTPException(404, "Scan not found")

    log.info("Received Chrome browser data for scan %s (%s)", scan_id, data.brand)

    # Convert Chrome data → unified signals dict
    signals = chrome_data_to_signals(data)

    # Score all 9 dimensions
    from .scanner.utils import SCORE_WEIGHTS
    scores = score_from_chrome_data(data, signals)

    total = round(sum(d["score"] for d in scores.values()), 1)
    from .scanner.engine import compute_grade, build_pitch_angles, _should_fallback_to_chrome
    grade = compute_grade(total)

    # Build pitch angles
    llm_probe = data.llm_probe
    pitch_angles = build_pitch_angles(
        brand_name=data.brand,
        scores=scores,
        detected_platform=data.homepage.detected_platform,
        sf_platform=run.sf_platform,
        platform_mismatch=bool(
            run.sf_platform and data.homepage.detected_platform and
            run.sf_platform.lower() != data.homepage.detected_platform.lower()
        ),
        llm_quote=llm_probe.quote_response,
        llm_failed=not llm_probe.can_quote,
        vertical=data.vertical_signals.detected_vertical,
        vertical_play="",
    )

    # Build recommendations
    from .scanner.utils import DIMENSION_LABELS, WHY_IT_MATTERS
    recommendations = [
        f"[{DIMENSION_LABELS.get(k, k)}] {v.get('finding', '')} — {WHY_IT_MATTERS.get(k, '')}"
        for k, v in sorted(scores.items(), key=lambda x: x[1].get("score", 0) / max(x[1].get("max_score", 1), 1))[:5]
    ]

    # Save screenshots to disk from base64
    from .scanner.browser import SS_BASE
    import re as _re
    safe_id = _re.sub(r"[^\w]", "_", scan_id)
    ss_dir = SS_BASE / safe_id
    ss_dir.mkdir(parents=True, exist_ok=True)
    saved_screenshots = []
    for label, b64_str in signals.get("screenshots_b64", {}).items():
        if not b64_str:
            continue
        # Strip data URI prefix if present
        if "base64," in b64_str:
            b64_str = b64_str.split("base64,", 1)[1]
        try:
            import base64 as _b64
            raw = _b64.b64decode(b64_str)
            ss_path = ss_dir / f"{label}.png"
            ss_path.write_bytes(raw)
            saved_screenshots.append(str(ss_path))
        except Exception as e:
            log.warning("Screenshot save failed [%s]: %s", label, e)

    # Generate PDF
    from .scanner.pdf_generator import generate as gen_pdf, render_html
    try:
        pdf_path = await gen_pdf(
            scan_id=scan_id,
            brand_name=data.brand,
            domain=run.domain,
            account_owner=run.triggered_by or "",
            overall_score=int(total),
            grade=grade,
            scores=scores,
            pitch_angles=pitch_angles,
            detected_platform=data.homepage.detected_platform,
            sf_platform=run.sf_platform,
            platform_mismatch=bool(
                run.sf_platform and data.homepage.detected_platform and
                run.sf_platform.lower() != data.homepage.detected_platform.lower()
            ),
            vertical=data.vertical_signals.detected_vertical,
            page_speed_score=data.page_speed.score,
            page_speed_lcp=f"{data.page_speed.lcp_ms}ms" if data.page_speed.lcp_ms else "",
            screenshot_paths=saved_screenshots,
            brand_logo_b64=await fetch_brand_logo(run.domain),
            scan_ts=data.audited_at or datetime.utcnow().isoformat(),
        )
    except Exception as e:
        log.error("PDF generation failed for Chrome data %s: %s", scan_id, e)
        pdf_path = None

    # Generate Slinger drafts
    from .scanner.slinger import build_context, generate_drafts
    ctx = build_context(
        brand_name=data.brand,
        domain=run.domain,
        overall_score=int(total),
        grade=grade,
        scores=scores,
        pitch_angles=pitch_angles,
        detected_platform=data.homepage.detected_platform,
        sf_platform=run.sf_platform,
        platform_mismatch=False,
        vertical=data.vertical_signals.detected_vertical,
        page_speed_score=data.page_speed.score,
        llm_quote=llm_probe.quote_response,
        llm_failed=not llm_probe.can_quote,
    )
    slinger = generate_drafts(data.brand, ctx, email_count=3)

    # Persist everything to DB
    now = datetime.utcnow()
    async with AsyncSessionLocal() as db:
        run = await db.get(ScanRun, scan_id)
        if run:
            run.status = "complete"
            run.scan_mode = "chrome"
            run.overall_score = int(total)
            run.grade = grade
            run.scores_json = scores
            run.signals_json = signals
            run.recommendations_json = recommendations
            run.pitch_angles_json = pitch_angles
            run.detected_platform = data.homepage.detected_platform
            run.platform_mismatch = bool(
                run.sf_platform and data.homepage.detected_platform and
                run.sf_platform.lower() != data.homepage.detected_platform.lower()
            )
            run.llm_probe_json = {
                "review_quote": llm_probe.quote_response,
                "complaint_quote": llm_probe.complaint_response,
                "failed": not llm_probe.can_quote,
            }
            run.pdf_path = pdf_path
            run.slinger_drafts_json = slinger
            run.screenshots_json = saved_screenshots
            run.chrome_raw_data = data.model_dump()
            run.chrome_job_status = "complete"
            run.chrome_job_completed_at = now
            run.chrome_pdps_visited = len(data.pdps_visited)
            # Append to existing audit_log
            existing_log = run.audit_log_json or []
            existing_log.append({
                "step": "chrome_data_received",
                "ts": now.isoformat(),
                "pdps_visited": len(data.pdps_visited),
                "reviews_found": len(signals.get("review_texts", [])),
                "audit_notes": data.audit_notes,
                "overall_score": int(total),
                "grade": grade,
            })
            run.audit_log_json = existing_log
            await db.commit()

        # Mark the chrome_job as complete
        from sqlalchemy import select
        job_result = await db.execute(
            select(ChromeJob).where(
                ChromeJob.scan_id == scan_id,
                ChromeJob.status == "running",
            ).limit(1)
        )
        job = job_result.scalar_one_or_none()
        if job:
            job.status = "complete"
            job.completed_at = now
            job.result_data = {"overall_score": int(total), "grade": grade}
            await db.commit()

    # Notify frontend via WebSocket
    async with AsyncSessionLocal() as db:
        run = await db.get(ScanRun, scan_id)
    if run:
        await ws_manager.broadcast(scan_id, {
            "type": "complete",
            "result": _serialize_run(run),
        })

    log.info(
        "Chrome audit complete for %s: %d/100 grade %s (%d PDPs, %d reviews)",
        data.brand, total, grade, len(data.pdps_visited), len(signals.get("review_texts", [])),
    )
    return {"status": "ok", "scan_id": scan_id, "score": int(total), "grade": grade}


@app.get("/api/scans/{scan_id}/screenshot/{label}")
async def api_screenshot(scan_id: str, label: str, user: dict = Depends(require_auth)):
    import re as _re
    # capture() sanitizes hyphens → underscores; match that here
    safe_id = _re.sub(r"[^\w]", "_", scan_id)
    ss_base = Path(os.environ.get("SCREENSHOTS_DIR", "/screenshots"))
    shot_path = ss_base / safe_id / f"{label}.png"
    if not shot_path.exists():
        raise HTTPException(404, "Screenshot not found")
    return FileResponse(str(shot_path), media_type="image/png")


@app.get("/api/check-recent")
async def api_check_recent(domain: str, user: dict = Depends(require_auth)):
    cutoff = datetime.utcnow() - timedelta(days=30)
    async with AsyncSessionLocal() as db:
        result = await db.execute(
            select(ScanRun)
            .where(ScanRun.domain == domain, ScanRun.triggered_at >= cutoff,
                   ScanRun.status == "complete")
            .order_by(desc(ScanRun.triggered_at))
            .limit(1)
        )
        run = result.scalar_one_or_none()
    if run:
        return {"found": True, "scan": _serialize_run(run)}
    return {"found": False}


@app.get("/api/scans/{scan_id}/export")
async def api_export_scan(scan_id: str, user: dict = Depends(require_auth)):
    """Export single scan as Excel."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    async with AsyncSessionLocal() as db:
        run = await db.get(ScanRun, scan_id)
    if not run:
        raise HTTPException(404, "Scan not found")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scan Results"
    _write_scan_to_sheet(ws, run)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="scan_{scan_id[:8]}.xlsx"'},
    )


@app.get("/api/history/export")
async def api_export_history(user: dict = Depends(require_auth)):
    """Export all completed scans as Excel (4-sheet format)."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from .scanner.utils import SCORE_WEIGHTS, DIMENSION_LABELS

    async with AsyncSessionLocal() as db:
        result = await db.execute(
            select(ScanRun).where(ScanRun.status == "complete")
            .order_by(desc(ScanRun.triggered_at))
        )
        runs = result.scalars().all()

    PURPLE = PatternFill("solid", fgColor="3C1053")
    WHITE_BOLD = Font(color="FFFFFF", bold=True, size=9)

    wb = openpyxl.Workbook()
    dim_keys = list(SCORE_WEIGHTS.keys())

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    headers = (["Brand", "Domain", "AE", "Score", "Grade", "Detected Platform",
                "SF Platform", "Mismatch", "Vertical", "Date", "Run By"]
               + [DIMENSION_LABELS[k] for k in dim_keys])
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.fill = PURPLE; c.font = WHITE_BOLD
    for r_idx, run in enumerate(runs, 2):
        scores = run.scores_json or {}
        row = [
            run.brand_name, run.domain, "", run.overall_score, run.grade,
            run.detected_platform or "", run.sf_platform or "",
            "YES" if run.platform_mismatch else "",
            (run.signals_json or {}).get("vertical", "") if run.signals_json else "",
            run.triggered_at.strftime("%Y-%m-%d") if run.triggered_at else "",
            run.triggered_by,
        ] + [scores.get(k, {}).get("score", 0) for k in dim_keys]
        for col, val in enumerate(row, 1):
            ws1.cell(row=r_idx, column=col, value=val)

    # Sheet 2: Pitch Angles
    ws2 = wb.create_sheet("Pitch Angles")
    for col, h in enumerate(["Brand", "Score", "Pitch 1", "Pitch 2", "Pitch 3"], 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = PURPLE; c.font = WHITE_BOLD
    sorted_runs = sorted(runs, key=lambda r: r.overall_score or 0)
    for r_idx, run in enumerate(sorted_runs, 2):
        pitches = (run.pitch_angles_json or []) + ["", "", ""]
        for col, val in enumerate([run.brand_name, run.overall_score] + pitches[:3], 1):
            c = ws2.cell(row=r_idx, column=col, value=val)
            if col >= 3:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    # Sheet 3: Recommendations
    ws3 = wb.create_sheet("Recommendations")
    for col, h in enumerate(["Brand", "Score", "Fix 1", "Fix 2", "Fix 3"], 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = PURPLE; c.font = WHITE_BOLD
    for r_idx, run in enumerate(sorted_runs, 2):
        recs = (run.recommendations_json or []) + ["", "", ""]
        for col, val in enumerate([run.brand_name, run.overall_score] + recs[:3], 1):
            c = ws3.cell(row=r_idx, column=col, value=val)
            if col >= 3:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    # Sheet 4: LLM Probe
    ws4 = wb.create_sheet("LLM Probe Results")
    for col, h in enumerate(["Brand", "LLM Score", "Review Quote", "Complaint", "Failed?"], 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.fill = PURPLE; c.font = WHITE_BOLD
    for r_idx, run in enumerate(runs, 2):
        llm = run.llm_probe_json or {}
        scores = run.scores_json or {}
        for col, val in enumerate([
            run.brand_name,
            scores.get("llm_crawlability", {}).get("score", 0),
            llm.get("review_quote", ""),
            llm.get("complaint_quote", ""),
            "YES" if llm.get("failed") else "",
        ], 1):
            c = ws4.cell(row=r_idx, column=col, value=val)
            if col in (3, 4):
                c.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="reviews_intelligence_report.xlsx"'},
    )


# ── WebSocket ─────────────────────────────────────────────────────────────────

@app.websocket("/ws/scans/{scan_id}")
async def ws_scan(scan_id: str, ws: WebSocket):
    if not ws.session.get("user"):
        await ws.close(code=1008)
        return

    await ws_manager.connect(scan_id, ws)
    try:
        # If scan already complete, send current state immediately
        async with AsyncSessionLocal() as db:
            run = await db.get(ScanRun, scan_id)
        if run and run.status in ("complete", "failed", "blocked"):
            await ws.send_json({"type": "status", "status": run.status, "result": _serialize_run(run)})
        # Keep connection alive until client disconnects
        while True:
            await ws.receive_text()
    except WebSocketDisconnect:
        ws_manager.disconnect(scan_id, ws)
    except Exception:
        ws_manager.disconnect(scan_id, ws)


# ── Frontend (serve React build) ──────────────────────────────────────────────

FRONTEND_DIST = Path(__file__).parent.parent / "frontend" / "dist"

if FRONTEND_DIST.exists():
    app.mount("/assets", StaticFiles(directory=str(FRONTEND_DIST / "assets")), name="assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_spa(request: Request, full_path: str):
        # Don't intercept API/auth/ws routes
        if full_path.startswith(("api/", "auth/", "ws/", "health")):
            raise HTTPException(404)
        index = FRONTEND_DIST / "index.html"
        if index.exists():
            return HTMLResponse(index.read_text())
        return HTMLResponse("<h1>Frontend not built. Run: cd frontend && npm run build</h1>")
else:
    @app.get("/", include_in_schema=False)
    async def root():
        return HTMLResponse(
            "<h1>Reviews Intelligence API</h1>"
            "<p>Frontend not built. Run: <code>cd frontend && npm run build</code></p>"
            "<p><a href='/docs'>API docs</a></p>"
        )


# ── Helpers ───────────────────────────────────────────────────────────────────

def _iso(dt) -> Optional[str]:
    """Serialize a DB datetime as UTC with an explicit offset.

    Timestamps are stored naive via datetime.utcnow(). A bare isoformat() has
    no timezone, so the browser's `new Date(...)` parses it as LOCAL time and
    every 'X min ago' goes negative. Stamping it as UTC fixes the whole UI.
    """
    if not dt:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.isoformat()


def _normalize_screenshots(raw) -> list:
    """Convert stored screenshots (list of path strings or dicts) to [{label, path}]."""
    if not raw:
        return []
    result = []
    for item in raw:
        if isinstance(item, str):
            label = Path(item).stem  # e.g. "homepage" from "/screenshots/.../homepage.png"
            result.append({"label": label, "path": item})
        elif isinstance(item, dict):
            result.append(item)
    return result


def _scan_evidence(run: ScanRun) -> dict:
    """Build the trust layer the UI uses to explain whether a scan is usable."""
    audit_log = run.audit_log_json or []
    scores = run.scores_json or {}
    screenshots = _normalize_screenshots(run.screenshots_json)
    chrome_raw = run.chrome_raw_data or {}

    pdps_found = len((chrome_raw.get("pdps_visited") or [])) if isinstance(chrome_raw, dict) else 0
    pdps_rendered = pdps_found
    reviews_found = 0
    schema_found = False
    category_stars = False

    if isinstance(chrome_raw, dict):
        for pdp in chrome_raw.get("pdps_visited") or []:
            reviews_found += len(pdp.get("reviews") or [])
            schema_found = schema_found or bool(
                pdp.get("has_review_schema") or pdp.get("has_aggregate_rating_schema")
            )
        category_stars = bool((chrome_raw.get("category_page") or {}).get("has_stars_on_cards"))

    for entry in audit_log:
        step = entry.get("step", "")
        if step == "pdp_discovery_result":
            pdps_found = max(pdps_found, int(entry.get("urls_found") or 0))
        elif step == "pdp_phase_complete":
            pdps_rendered = max(pdps_rendered, int(entry.get("pdps_rendered") or 0))
        elif step.startswith("pdp_review_audit_"):
            reviews_found += int(entry.get("review_texts_found") or 0)
        elif step == "chrome_data_received":
            pdps_rendered = max(pdps_rendered, int(entry.get("pdps_visited") or 0))
            reviews_found = max(reviews_found, int(entry.get("reviews_found") or 0))

    for value in scores.values():
        finding = (value or {}).get("finding", "").lower()
        schema_found = schema_found or "schema" in finding and "no " not in finding[:12]
        category_stars = category_stars or "category" in finding and "star" in finding and "no " not in finding[:12]

    score = 25
    if run.scan_mode == "chrome":
        score += 30
    if pdps_found:
        score += 12
    if pdps_rendered:
        score += 12
    if reviews_found:
        score += 18
    if screenshots:
        score += 8
    if schema_found:
        score += 5

    zero_count = sum(
        1
        for key, value in scores.items()
        if key != "llm_crawlability" and (value or {}).get("score") == 0
    )
    if run.status == "blocked":
        score = 0
        level = "blocked"
        summary = "The site blocked the scanner. Do not use the score until Browser Scan succeeds."
    elif run.chrome_job_status in ("queued", "running"):
        score = 45
        level = "pending"
        summary = "Browser verification is queued or running. Treat the current score as provisional."
    else:
        score = max(0, min(100, score - min(25, zero_count * 5)))
        if score >= 80:
            level = "high"
            summary = "Strong evidence: the score is backed by rendered PDPs, reviews, and screenshots."
        elif score >= 55:
            level = "medium"
            summary = "Usable brief, but a Browser Scan or more review evidence would strengthen it."
        else:
            level = "low"
            summary = "Low confidence: the scan found limited evidence, so verify before outreach."

    proof = [
        f"{pdps_found} product pages discovered",
        f"{pdps_rendered} product pages rendered",
        f"{reviews_found} review texts extracted",
        f"{len(screenshots)} screenshots captured",
    ]
    if schema_found:
        proof.append("Review schema evidence found")
    if category_stars:
        proof.append("Category-page star signal found")

    gaps = []
    if pdps_found == 0:
        gaps.append("No PDPs discovered")
    if reviews_found == 0:
        gaps.append("No review text extracted")
    if not screenshots:
        gaps.append("No screenshots captured")
    if not schema_found:
        gaps.append("No review schema evidence")
    if not category_stars:
        gaps.append("No category star evidence")

    if level in {"blocked", "low", "pending"}:
        action = "Run or wait for Browser Scan before using this account brief."
    elif run.platform_mismatch:
        action = "Lead with the platform mismatch, then use the weakest dimensions as proof."
    else:
        action = "Use the top pitch angles and screenshots as the outreach spine."

    return {
        "level": level,
        "score": score,
        "summary": summary,
        "proof": proof,
        "gaps": gaps[:5],
        "next_action": action,
    }


def _serialize_run(run: ScanRun) -> dict:
    return {
        "id": str(run.id),
        "brand_name": run.brand_name,
        "domain": run.domain,
        "triggered_by": run.triggered_by,
        "triggered_at": _iso(run.triggered_at),
        "status": run.status,
        "overall_score": run.overall_score,
        "grade": run.grade,
        "scores": run.scores_json,
        "recommendations": run.recommendations_json,
        "pitch_angles": run.pitch_angles_json,
        "llm_probe": run.llm_probe_json,
        "detected_platform": run.detected_platform,
        "sf_platform": run.sf_platform,
        "platform_mismatch": run.platform_mismatch,
        "pdf_path": run.pdf_path,
        "slinger_drafts": run.slinger_drafts_json,
        "screenshots": _normalize_screenshots(run.screenshots_json),
        "error_message": run.error_message,
        "audit_log": run.audit_log_json or [],
        "scan_mode": run.scan_mode or "playwright",
        "chrome_job_status": run.chrome_job_status,
        "scan_fallback_reason": run.scan_fallback_reason,
        "chrome_pdps_visited": run.chrome_pdps_visited or 0,
        "evidence": _scan_evidence(run),
    }


def _write_scan_to_sheet(ws, run: ScanRun):
    from .scanner.utils import DIMENSION_LABELS, SCORE_WEIGHTS
    ws.append(["Brand", run.brand_name])
    ws.append(["Domain", run.domain])
    ws.append(["Score", run.overall_score])
    ws.append(["Grade", run.grade])
    ws.append(["Platform", run.detected_platform or ""])
    ws.append([])
    ws.append(["Dimension", "Score", "Max", "Finding"])
    scores = run.scores_json or {}
    for key in SCORE_WEIGHTS:
        dim = scores.get(key, {})
        ws.append([
            DIMENSION_LABELS.get(key, key),
            dim.get("score", 0),
            dim.get("max_score", SCORE_WEIGHTS[key]),
            dim.get("finding", ""),
        ])
